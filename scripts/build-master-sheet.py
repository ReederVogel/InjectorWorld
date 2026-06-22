"""
Build injectors.world master import sheet.
Run: python scripts/build-master-sheet.py
Output: data/injectors-world-master.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── palette ────────────────────────────────────────────────────────────────────
NAVY        = "0B1B34"
MINT        = "3FA68A"
MINT_SOFT   = "E6F2EE"
RED_HDR     = "C0392B"   # required column header bg
BLUE_HDR    = "1E3A5F"   # optional column header bg
GRAY_ROW    = "F2F2F2"   # example row bg
WHITE       = "FFFFFF"
LIGHT_MINT  = "D6EDE8"
WARN_YELLOW = "FFF3CD"

FONT_NAME   = "Arial"

def hdr_font(color=WHITE):
    return Font(name=FONT_NAME, bold=True, color=color, size=10)

def body_font(bold=False, italic=False, color="000000"):
    return Font(name=FONT_NAME, bold=bold, italic=italic, color=color, size=10)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_wrap():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, font=None, fill_=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font    = font
    if fill_:  c.fill   = fill_
    if align:  c.alignment = align
    if border: c.border = border
    return c

# ── field definitions ───────────────────────────────────────────────────────────
# (name, required, width, note)
CLINIC_FIELDS = [
    ("clinic_id",             True,  22, "Unique ID, e.g. clinic-rs-123456"),
    ("clinic_name",           True,  30, "Full practice name"),
    ("address_line_1",        True,  30, "Street address"),
    ("address_line_2",        False, 18, "Suite / floor"),
    ("city",                  True,  18, "City name, e.g. Scottsdale"),
    ("state",                 True,  10, "2-letter, e.g. AZ"),
    ("zip",                   True,  10, "5-digit ZIP"),
    ("neighborhood",          False, 20, "e.g. Upper East Side (critical for SEO)"),
    ("county",                False, 18, "e.g. Maricopa County"),
    ("country",               False, 10, "Default: US"),
    ("latitude",              True,  14, "6 decimal places"),
    ("longitude",             True,  14, "6 decimal places"),
    ("google_place_id",       False, 28, "From Google Places API"),
    ("google_maps_url",       False, 40, "Direct Google Maps link"),
    ("directions_url",        False, 40, "Pre-built directions URL"),
    ("apple_maps_url",        False, 40, "For iOS users"),
    ("phone",                 False, 18, "E.164 format: +12025550100"),
    ("email",                 False, 25, "info@clinic.com"),
    ("website_url",           False, 35, "https://..."),
    ("booking_url",           False, 35, "Direct booking link"),
    ("hours_json",            False, 40, '{"mon":"9-17","tue":"9-17","sun":"closed"}'),
    ("service_type",          False, 14, "In-Person / Telehealth / Both"),
    ("accepts_insurance",     False, 14, "true / false"),
    ("payment_methods",       False, 30, "Visa; Mastercard; CareCredit"),
    ("amenities",             False, 30, "Parking; Wheelchair accessible"),
    ("logo_url",              False, 35, "Logo image URL"),
    ("clinic_photo_urls",     False, 40, "url1; url2; url3 (semicolon-separated)"),
    ("aggregate_rating",      False, 14, "e.g. 4.8"),
    ("aggregate_rating_count",False, 14, "Total review count"),
    ("year_established",      False, 14, "e.g. 2015"),
    ("provider_ids",          False, 40, "prov-rs-001; prov-rs-002 (semicolon-sep)"),
    ("source_urls",           True,  40, "All URLs scraped from (semicolon-sep)"),
    ("last_scraped_date",     True,  16, "YYYY-MM-DD"),
]

PROVIDER_FIELDS = [
    ("provider_id",              True,  24, "Unique ID, e.g. prov-rs-arizona-001"),
    ("full_name",                True,  22, "First + Last only, no titles"),
    ("credentials",              True,  14, "MD / DO / NP / PA / RN / DDS"),
    ("title",                    True,  30, "e.g. Board-Certified Dermatologist"),
    ("board_certifications",     False, 30, "Dermatology; AAD Fellow (semicolon-sep)"),
    ("license_number",           False, 18, "From state medical board"),
    ("license_state",            False, 10, "2-letter, e.g. AZ"),
    ("license_status",           False, 12, "Active / Inactive / Expired"),
    ("license_verification_url", False, 40, "Direct link to state board record"),
    ("npi_number",               False, 14, "10-digit NPI"),
    ("years_experience",         False, 14, "e.g. 12"),
    ("year_started_practicing",  False, 14, "e.g. 2012"),
    ("clinic_id",                True,  24, "Must match a clinic_id in Clinics tab"),
    ("additional_clinic_ids",    False, 40, "Secondary clinics (semicolon-sep)"),
    ("tagline",                  False, 40, "Max 100 chars, short pitch"),
    ("bio",                      False, 50, "2-3 paragraph bio"),
    ("profile_photo_url",        False, 40, "Headshot URL"),
    ("languages",                False, 24, "English; Spanish (semicolon-sep)"),
    ("gender",                   False, 12, "Female / Male / Non-binary / Unknown"),
    ("treatments_offered",       True,  40, "Botox; Lip Filler; Sculptra (semicolon-sep)"),
    ("specialties",              False, 40, "Forehead; Crow's Feet (semicolon-sep)"),
    ("pricing_botox_per_unit",   False, 14, "USD per unit, e.g. 16"),
    ("pricing_filler_per_syringe",False,14, "USD per syringe, e.g. 850"),
    ("pricing_consultation",     False, 14, "USD, 0 if free"),
    ("accepts_new_patients",     False, 14, "true / false"),
    ("offers_virtual_consult",   False, 14, "true / false"),
    ("offers_in_person",         False, 14, "true / false"),
    ("website_url",              False, 35, "https://..."),
    ("email",                    False, 25, "provider@clinic.com"),
    ("phone_direct",             False, 18, "E.164 format: +12025550100"),
    ("instagram_url",            False, 35, "https://instagram.com/..."),
    ("tiktok_url",               False, 35, "https://tiktok.com/..."),
    ("linkedin_url",             False, 35, "https://linkedin.com/..."),
    ("aggregate_rating",         False, 14, "e.g. 4.9"),
    ("aggregate_rating_count",   False, 14, "Total review count"),
    ("source_urls",              True,  40, "All URLs scraped from (semicolon-sep)"),
    ("last_scraped_date",        True,  16, "YYYY-MM-DD"),
]

REVIEW_FIELDS = [
    ("review_id",            True,  20, "Unique ID, e.g. rev-rs-001"),
    ("clinic_id",            True,  24, "Must match a clinic_id in Clinics tab"),
    ("provider_id",          False, 24, "Optional — links to a specific provider"),
    ("reviewer_first_name",  False, 16, "First name only"),
    ("reviewer_initial",     False, 10, "Last initial, e.g. R"),
    ("reviewer_age_range",   False, 14, "e.g. 30-34"),
    ("reviewer_city",        False, 16, "e.g. Brooklyn"),
    ("rating",               True,  10, "Integer 1-5"),
    ("review_title",         False, 30, "e.g. Best Botox in NYC"),
    ("review_text",          True,  50, "Full review body"),
    ("treatment_tag",        False, 20, "e.g. Botox"),
    ("review_date",          True,  14, "YYYY-MM-DD"),
    ("source_platform",      True,  16, "google/yelp/healthgrades/vitals/zocdoc/clinic_site"),
    ("source_url",           True,  40, "Original review URL"),
    ("response_from_provider",False,50, "Provider reply text"),
    ("response_date",        False, 14, "YYYY-MM-DD"),
]

PHOTO_FIELDS = [
    ("photo_id",             True,  20, "Unique ID, e.g. ph-001"),
    ("provider_id",          False, 24, "Optional — photo owner"),
    ("clinic_id",            False, 24, "Optional — photo owner"),
    ("photo_url",            True,  40, "Direct image URL"),
    ("type",                 True,  20, "before/after/headshot/clinic_interior/clinic_exterior/treatment_room/team/equipment"),
    ("pair_id",              False, 16, "Links before + after pair, e.g. pair-001"),
    ("weeks_post_treatment", False, 16, "Integer, only for 'after' photos"),
    ("treatment_tag",        False, 20, "e.g. Lip Filler"),
    ("caption",              False, 40, "Photo caption"),
    ("consent_documented",   False, 14, "true / false — MUST be true before going live"),
    ("source_platform",      True,  20, "clinic_site / google / instagram"),
    ("source_url",           True,  40, "Source page URL"),
]

US_STATES = [
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA",
    "HI","ID","IL","IN","IA","KS","KY","LA","ME","MD",
    "MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
    "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC",
    "SD","TN","TX","UT","VT","VA","WA","WV","WI","WY","DC",
]

TREATMENTS = [
    "Botox","Dysport","Xeomin","Jeuveau","Daxxify",
    "Lip Filler","Cheek Filler","Tear Trough","Masseter Botox",
    "Sculptra","Kybella","PRP","Microneedling","Thread Lift",
]

def build_instructions(wb):
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 70

    ws.row_dimensions[1].height = 60
    ws.merge_cells("B1:C1")
    c = ws["B1"]
    c.value = "injector.world — Master Import Sheet"
    c.font = Font(name=FONT_NAME, bold=True, size=18, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = center()

    ws.row_dimensions[2].height = 14

    rows = [
        ("HOW TO USE", None, True),
        ("1. Clinics tab", "Fill in every clinic first. Each row = one location.", False),
        ("2. Providers tab", "Each row = one injector. clinic_id must match a value from the Clinics tab.", False),
        ("3. Reviews tab", "Each row = one review. clinic_id must match Clinics tab.", False),
        ("4. Photos tab", "Each row = one photo. provider_id or clinic_id must match.", False),
        ("5. Convert & upload", "Run:  python scripts/excel-to-import.py\nThen upload data/import/combined.csv via Admin → Import.", False),
        ("", None, False),
        ("COLOR CODING", None, True),
        ("Red header", "Required field — cannot be blank. Row will be skipped on import.", False),
        ("Blue header", "Optional field — leave blank if not available.", False),
        ("Gray example row", "Row 2 on each tab is a sample. Delete or overwrite before uploading.", False),
        ("", None, False),
        ("IMPORTANT RULES", None, True),
        ("clinic_id / provider_id", "Keep the scraper IDs exactly as-is (e.g. clinic-rs-123456). Do not renumber.", False),
        ("treatments_offered", "Semicolon-separated. Use exact names: Botox; Lip Filler; Sculptra", False),
        ("provider_ids (on clinic)", "Semicolon-separated list of all provider_ids working at this clinic.", False),
        ("Dates", "Always YYYY-MM-DD format, e.g. 2026-06-22", False),
        ("Phones", "E.164 format: +12025550100 (no dashes or spaces)", False),
        ("source_urls", "At least one URL per clinic/provider. Audit trail.", False),
        ("consent_documented", "Before/After photos MUST have true before going live.", False),
        ("", None, False),
        ("VALID TREATMENTS", "Botox; Dysport; Xeomin; Jeuveau; Daxxify; Lip Filler; Cheek Filler; Tear Trough; Masseter Botox; Sculptra; Kybella; PRP; Microneedling; Thread Lift", False),
        ("VALID CREDENTIALS", "MD, DO, NP, PA, RN, DDS", False),
        ("VALID SOURCE PLATFORMS", "google, yelp, healthgrades, vitals, zocdoc, clinic_site", False),
        ("VALID PHOTO TYPES", "before, after, headshot, clinic_interior, clinic_exterior, treatment_room, team, equipment", False),
    ]

    r = 3
    for label, desc, is_section in rows:
        ws.row_dimensions[r].height = 28 if is_section else 22
        if is_section:
            ws.merge_cells(f"B{r}:C{r}")
            c = ws.cell(row=r, column=2, value=label)
            c.font = Font(name=FONT_NAME, bold=True, color=WHITE, size=11)
            c.fill = fill(MINT)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        elif label == "":
            ws.row_dimensions[r].height = 8
        else:
            ws.cell(row=r, column=2, value=label).font = Font(name=FONT_NAME, bold=True, size=10)
            c = ws.cell(row=r, column=3, value=desc)
            c.font = body_font()
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    ws.freeze_panes = None
    return ws


def build_tab(wb, title, fields, example_row):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False

    # ── row 1: tab title banner ────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    ws.merge_cells(f"A1:{get_column_letter(len(fields))}1")
    c = ws["A1"]
    c.value = f"injector.world — {title}"
    c.font = Font(name=FONT_NAME, bold=True, size=13, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = center()

    # ── row 2: field name headers ──────────────────────────────────────────
    ws.row_dimensions[2].height = 40
    for col_idx, (fname, required, width, _) in enumerate(fields, start=1):
        bg = RED_HDR if required else BLUE_HDR
        marker = " *" if required else ""
        c = ws.cell(row=2, column=col_idx, value=fname + marker)
        c.font = hdr_font()
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── row 3: field notes (light mint bg) ────────────────────────────────
    ws.row_dimensions[3].height = 36
    for col_idx, (_, _, _, note) in enumerate(fields, start=1):
        c = ws.cell(row=3, column=col_idx, value=note)
        c.font = Font(name=FONT_NAME, italic=True, size=9, color="555555")
        c.fill = fill("EAF4F1")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = thin_border()

    # ── row 4: example row (gray) ──────────────────────────────────────────
    ws.row_dimensions[4].height = 22
    for col_idx, val in enumerate(example_row, start=1):
        c = ws.cell(row=4, column=col_idx, value=val)
        c.font = Font(name=FONT_NAME, italic=True, size=9, color="888888")
        c.fill = fill(GRAY_ROW)
        c.alignment = left_wrap()
        c.border = thin_border()

    # ── row 5+: data area (alternating white/very light) ──────────────────
    for row_idx in range(5, 205):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(fields) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = fill(WHITE) if row_idx % 2 == 1 else fill("F9FBFA")
            c.font = body_font()
            c.alignment = left_wrap()
            c.border = thin_border()

    ws.freeze_panes = "A5"
    return ws


def add_dropdown(ws, col_idx, formula, first_row=5, last_row=204, prompt="", title=""):
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error=f"Please select from the dropdown list.",
        showInputMessage=bool(prompt),
        promptTitle=title or None,
        prompt=prompt or None,
    )
    col = get_column_letter(col_idx)
    dv.sqref = f"{col}{first_row}:{col}{last_row}"
    ws.add_data_validation(dv)


def col_for(fields, name):
    for i, (fname, *_) in enumerate(fields, start=1):
        if fname == name:
            return i
    return None


def build_clinics(wb):
    example = [
        "clinic-rs-123456", "Infini Cosmetic Associates", "8841 E Bell Rd",
        "Suite 105", "Scottsdale", "AZ", "85260", "North Scottsdale", "Maricopa",
        "US", "33.623400", "-111.891200", "", "", "", "",
        "+14805550100", "info@infini.com", "https://infinicosmetic.com", "",
        '{"mon":"9-17","tue":"9-17","wed":"9-17","thu":"9-17","fri":"9-17","sat":"closed","sun":"closed"}',
        "In-Person", "false", "Visa; Mastercard; CareCredit", "Parking; Wheelchair accessible",
        "", "", "4.9", "312", "2011",
        "prov-rs-arizona-001; prov-rs-arizona-002",
        "https://realself.com/find/arizona/scottsdale/infini-cosmetic",
        "2026-06-22",
    ]
    ws = build_tab(wb, "Clinics", CLINIC_FIELDS, example)

    # Dropdowns
    state_col  = col_for(CLINIC_FIELDS, "state")
    stype_col  = col_for(CLINIC_FIELDS, "service_type")
    ins_col    = col_for(CLINIC_FIELDS, "accepts_insurance")

    state_list = '","'.join(US_STATES)
    add_dropdown(ws, state_col,  f'"{state_list}"', prompt="2-letter US state code", title="State")
    add_dropdown(ws, stype_col,  '"In-Person,Telehealth,Both"', prompt="Select service type", title="Service Type")
    add_dropdown(ws, ins_col,    '"true,false"', prompt="Does this clinic accept insurance?", title="Accepts Insurance")


def build_providers(wb):
    example = [
        "prov-rs-arizona-001", "Sarah Kim", "MD",
        "Board-Certified Plastic Surgeon", "Plastic Surgery; ASPS Member",
        "271832", "AZ", "Active", "https://azmd.gov/lookup?id=271832",
        "1234567890", "14", "2012",
        "clinic-rs-123456", "",
        "Conservative, natural-looking results", "",
        "https://infinicosmetic.com/team/sarah-kim", "English; Korean",
        "Female", "Botox; Lip Filler; Cheek Filler; Sculptra", "Forehead; Tear Trough",
        "16", "850", "100",
        "true", "false", "true",
        "https://infinicosmetic.com", "dr.kim@infini.com", "+14805550101",
        "", "", "",
        "4.9", "187",
        "https://realself.com/find/arizona/scottsdale/sarah-kim-md",
        "2026-06-22",
    ]
    ws = build_tab(wb, "Providers", PROVIDER_FIELDS, example)

    cred_col    = col_for(PROVIDER_FIELDS, "credentials")
    lic_st_col  = col_for(PROVIDER_FIELDS, "license_state")
    lic_s_col   = col_for(PROVIDER_FIELDS, "license_status")
    gender_col  = col_for(PROVIDER_FIELDS, "gender")
    anp_col     = col_for(PROVIDER_FIELDS, "accepts_new_patients")
    vc_col      = col_for(PROVIDER_FIELDS, "offers_virtual_consult")
    ip_col      = col_for(PROVIDER_FIELDS, "offers_in_person")

    add_dropdown(ws, cred_col,   '"MD,DO,NP,PA,RN,DDS"', title="Credentials")
    state_list = '","'.join(US_STATES)
    add_dropdown(ws, lic_st_col, f'"{state_list}"', title="License State")
    add_dropdown(ws, lic_s_col,  '"Active,Inactive,Expired"', title="License Status")
    add_dropdown(ws, gender_col, '"Female,Male,Non-binary,Unknown"', title="Gender")
    add_dropdown(ws, anp_col,    '"true,false"', title="Accepts New Patients")
    add_dropdown(ws, vc_col,     '"true,false"', title="Offers Virtual Consult")
    add_dropdown(ws, ip_col,     '"true,false"', title="Offers In-Person")


def build_reviews(wb):
    example = [
        "rev-rs-001", "clinic-rs-123456", "prov-rs-arizona-001",
        "Maya", "R", "30-34", "Scottsdale",
        "5", "Incredible results, so natural",
        "Dr. Kim was amazing. My lips look completely natural and exactly what I asked for.",
        "Lip Filler", "2026-05-14",
        "google", "https://google.com/maps/place/.../reviews/...",
        "", "",
    ]
    ws = build_tab(wb, "Reviews", REVIEW_FIELDS, example)

    platform_col = col_for(REVIEW_FIELDS, "source_platform")
    rating_col   = col_for(REVIEW_FIELDS, "rating")

    add_dropdown(ws, platform_col, '"google,yelp,healthgrades,vitals,zocdoc,clinic_site"', title="Source Platform")

    dv = DataValidation(
        type="whole", operator="between",
        formula1="1", formula2="5",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid rating",
        error="Rating must be a whole number between 1 and 5.",
    )
    rating_letter = get_column_letter(rating_col)
    dv.sqref = f"{rating_letter}5:{rating_letter}204"
    ws.add_data_validation(dv)


def build_photos(wb):
    example = [
        "ph-001", "prov-rs-arizona-001", "clinic-rs-123456",
        "https://infinicosmetic.com/photos/sarah-kim-headshot.jpg",
        "headshot", "", "", "",
        "Dr. Kim professional headshot", "true",
        "clinic_site", "https://infinicosmetic.com/team/sarah-kim",
    ]
    ws = build_tab(wb, "Photos", PHOTO_FIELDS, example)

    type_col     = col_for(PHOTO_FIELDS, "type")
    consent_col  = col_for(PHOTO_FIELDS, "consent_documented")
    platform_col = col_for(PHOTO_FIELDS, "source_platform")

    add_dropdown(ws, type_col,     '"before,after,headshot,clinic_interior,clinic_exterior,treatment_room,team,equipment"', title="Photo Type")
    add_dropdown(ws, consent_col,  '"true,false"', title="Consent Documented")
    add_dropdown(ws, platform_col, '"clinic_site,google,instagram,realself"', title="Source Platform")


def main():
    import os
    os.makedirs("data/import", exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_instructions(wb)
    build_clinics(wb)
    build_providers(wb)
    build_reviews(wb)
    build_photos(wb)

    out = "data/injectors-world-master.xlsx"
    wb.save(out)
    print(f"[OK] Saved: {out}")
    print("  Tabs: Instructions | Clinics | Providers | Reviews | Photos")
    print("  Row 2: column headers (red=required, blue=optional)")
    print("  Row 3: field notes")
    print("  Row 4: example row (delete before uploading)")
    print("  Row 5+: your data")
    print()
    print("Next: python scripts/excel-to-import.py")


if __name__ == "__main__":
    main()
