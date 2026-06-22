"""
Convert the filled master Excel sheet to a combined import CSV.

Usage:
  python scripts/excel-to-import.py
  python scripts/excel-to-import.py --dry-run        (validate only, no output file)
  python scripts/excel-to-import.py --input path/to/custom.xlsx

Output: data/import/combined.csv  (upload this via Admin → Import CSV)
        data/import/import-report.txt
"""

import argparse
import csv
import os
import sys
from datetime import date

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

SKIP_ROWS = {1, 2, 3, 4}  # banner, header, notes, example

# Maps tab name → (record_type, required columns)
TAB_CONFIG = {
    "Clinics":   ("clinic",   ["clinic_id", "clinic_name", "latitude", "longitude", "city", "state", "source_urls", "last_scraped_date"]),
    "Providers": ("provider", ["provider_id", "full_name", "credentials", "clinic_id", "treatments_offered", "source_urls", "last_scraped_date"]),
    "Reviews":   ("review",   ["review_id", "clinic_id", "rating", "review_text", "review_date", "source_platform", "source_url"]),
    "Photos":    ("photo",    ["photo_id", "photo_url", "type", "source_platform", "source_url"]),
}


def cell_val(cell):
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def read_tab(ws):
    """Read a tab, returning (headers, rows_as_dicts). Skips banner/notes/example rows."""
    headers = []
    rows = []

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx == 2:
            # Header row — strip the " *" required marker
            for cell in row:
                h = cell_val(cell).rstrip(" *")
                headers.append(h)
            continue
        if row_idx in SKIP_ROWS:
            continue
        vals = [cell_val(c) for c in row]
        # Skip fully blank rows
        if not any(v for v in vals):
            continue
        row_dict = dict(zip(headers, vals))
        rows.append(row_dict)

    return headers, rows


def validate_row(row_dict, required_cols, record_type, row_num, errors):
    ok = True
    for col in required_cols:
        if not row_dict.get(col, "").strip():
            errors.append(f"  [{record_type.upper()} row {row_num}] Missing required field: {col}")
            ok = False
    return ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",   default="data/injectors-world-master.xlsx")
    parser.add_argument("--output",  default="data/import/combined.csv")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: File not found: {args.input}")
        print("       Build the template first: python scripts/build-master-sheet.py")
        sys.exit(1)

    print(f"Reading: {args.input}")
    wb = openpyxl.load_workbook(args.input, data_only=True)

    all_rows = []
    all_headers = set()
    counts = {}
    errors = []
    warnings = []

    for tab_name, (record_type, required_cols) in TAB_CONFIG.items():
        if tab_name not in wb.sheetnames:
            warnings.append(f"Tab '{tab_name}' not found — skipped.")
            continue

        ws = wb[tab_name]
        headers, rows = read_tab(ws)

        if not headers:
            warnings.append(f"Tab '{tab_name}' has no header row — skipped.")
            continue

        valid_count = 0
        skip_count = 0

        for i, row_dict in enumerate(rows, start=5):  # data starts at row 5
            ok = validate_row(row_dict, required_cols, record_type, i, errors)
            if ok:
                row_dict["record_type"] = record_type
                all_rows.append(row_dict)
                all_headers.update(row_dict.keys())
                valid_count += 1
            else:
                skip_count += 1

        counts[tab_name] = {"valid": valid_count, "skipped": skip_count, "record_type": record_type}
        print(f"  {tab_name}: {valid_count} valid, {skip_count} skipped")

    # ── report ───────────────────────────────────────────────────────────────
    report_lines = [
        f"injectors.world Import Conversion Report",
        f"Generated: {date.today().isoformat()}",
        f"Source:    {args.input}",
        f"",
        f"{'Tab':<14} {'Record type':<12} {'Valid':>7} {'Skipped':>8}",
        f"{'-'*46}",
    ]
    total_valid = 0
    total_skip = 0
    for tab, c in counts.items():
        report_lines.append(f"{tab:<14} {c['record_type']:<12} {c['valid']:>7} {c['skipped']:>8}")
        total_valid += c["valid"]
        total_skip  += c["skipped"]
    report_lines += [
        f"{'-'*46}",
        f"{'TOTAL':<14} {'':<12} {total_valid:>7} {total_skip:>8}",
        "",
    ]

    if warnings:
        report_lines.append("WARNINGS:")
        for w in warnings:
            report_lines.append(f"  {w}")
        report_lines.append("")

    if errors:
        report_lines.append(f"VALIDATION ERRORS ({len(errors)} rows skipped):")
        for e in errors:
            report_lines.append(e)
        report_lines.append("")

    if not errors:
        report_lines.append("No validation errors. All rows are ready to upload.")

    report_text = "\n".join(report_lines)
    print()
    print(report_text)

    # ── dry run stops here ───────────────────────────────────────────────────
    if args.dry_run:
        print("\nDry run — no output file written.")
        if errors:
            sys.exit(1)
        return

    if not all_rows:
        print("No valid rows to write.")
        sys.exit(1)

    # ── write combined CSV ───────────────────────────────────────────────────
    os.makedirs(os.path.dirname(args.output), exist_ok=True)

    # Column order: record_type first, then all others sorted
    all_headers.discard("record_type")
    fieldnames = ["record_type"] + sorted(all_headers)

    with open(args.output, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in all_rows:
            writer.writerow(row)

    print(f"\nOutput: {args.output}  ({total_valid} rows)")

    # Save report
    report_path = args.output.replace("combined.csv", "import-report.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_text)
    print(f"Report: {report_path}")

    if errors:
        print(f"\n{len(errors)} rows were skipped due to missing required fields.")
        print("Fix them in the Excel sheet and re-run this script.")
        sys.exit(1)
    else:
        print("\nReady. Upload data/import/combined.csv via Admin → Import CSV.")


if __name__ == "__main__":
    main()
